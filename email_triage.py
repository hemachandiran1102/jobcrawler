#!/usr/bin/env python3
"""
email_triage.py — Multi-Account & Multi-Lingual Email Triage Pipeline
====================================================================
Connects securely to multiple email inboxes simultaneously (1 Gmail + multiple Outlook/IMAP),
scans recruiter & hiring platform responses in any language (English, German, French, Dutch,
Nordic languages, Arabic, Spanish, Portuguese, Italian, Polish), and extracts actionable
"Proceed to Next Steps" responses (interview invites, screening calls, technical assessments, availability requests).

Supported Providers:
  - Gmail (Google App Password)
  - Microsoft Outlook / Office 365 / Hotmail (Microsoft App Password)
  - Custom IMAP / Yahoo / iCloud

Outputs:
  - Master Excel: interview_pipeline.xlsx (Action Required + Per-Category Tabs + Account breakdown)
  - Master CSV  : interview_pipeline.csv
  - Master JSON : interview_pipeline.json (for web dashboard)
"""

import os
import re
import sys
import json
import html as html_module
import time
import email
import email.message
import email.utils
from email.header import decode_header
import imaplib
import argparse
import urllib.parse
from datetime import datetime, timedelta
from pathlib import Path

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


# ══════════════════════════════════════════════════════════════════════
# PATHS & CONSTANTS
# ══════════════════════════════════════════════════════════════════════
_SCRIPT_DIR = Path(__file__).parent.resolve()
WORK_DIR = Path(os.environ.get("WORK_DIR", str(_SCRIPT_DIR)))
CONFIG_PATH = WORK_DIR / "email_config.json"
ENV_PATH = WORK_DIR / ".env"

MASTER_EXCEL_PATH = WORK_DIR / "interview_pipeline.xlsx"
MASTER_CSV_PATH   = WORK_DIR / "interview_pipeline.csv"
MASTER_JSON_PATH  = WORK_DIR / "interview_pipeline.json"

# Scheduling & Meeting link domains
MEETING_LINK_PATTERNS = [
    r'https?://(?:www\.)?calendly\.com/[^\s"\'<>]+',
    r'https?://(?:www\.)?goodtime\.io/[^\s"\'<>]+',
    r'https?://(?:www\.)?savvycal\.com/[^\s"\'<>]+',
    r'https?://(?:www\.)?chilipiper\.com/[^\s"\'<>]+',
    r'https?://(?:www\.)?hubspot\.com/meetings/[^\s"\'<>]+',
    r'https?://(?:www\.)?tidycal\.com/[^\s"\'<>]+',
    r'https?://(?:www\.)?youcanbook\.me/[^\s"\'<>]+',
    r'https?://(?:[a-zA-Z0-9-]+\.)?greenhouse\.io/[^\s"\'<>]+',
    r'https?://(?:[a-zA-Z0-9-]+\.)?lever\.co/[^\s"\'<>]+',
    r'https?://(?:[a-zA-Z0-9-]+\.)?smartrecruiters\.com/[^\s"\'<>]+',
    r'https?://(?:[a-zA-Z0-9-]+\.)?workday\.com/[^\s"\'<>]+',
    r'https?://(?:[a-zA-Z0-9-]+\.)?ashbyhq\.com/[^\s"\'<>]+',
    r'https?://(?:www\.)?hackerrank\.com/[^\s"\'<>]+',
    r'https?://(?:www\.)?codility\.com/[^\s"\'<>]+',
    r'https?://(?:www\.)?testgorilla\.com/[^\s"\'<>]+',
    r'https?://(?:www\.)?hirevue\.com/[^\s"\'<>]+',
    r'https?://meet\.google\.com/[a-z]{3}-[a-z]{4}-[a-z]{3}',
    r'https?://(?:[a-zA-Z0-9-]+\.)?zoom\.us/j/[0-9]+[^\s"\'<>]*',
    r'https?://teams\.microsoft\.com/l/meetup-join/[^\s"\'<>]+',
]

KNOWN_TECH_ROLES = [
    "DevOps Engineer", "Senior DevOps Engineer", "Lead DevOps Engineer",
    "Cloud Engineer", "Senior Cloud Engineer", "Cloud Infrastructure Engineer", "Cloud Architect",
    "Site Reliability Engineer", "Senior SRE", "SRE Lead",
    "Platform Engineer", "Senior Platform Engineer", "Staff Platform Engineer",
    "Infrastructure Engineer", "Senior Infrastructure Engineer",
    "AWS DevOps Engineer", "Azure DevOps Engineer", "Kubernetes Engineer",
    "Software Engineer", "Senior Software Engineer", "Full Stack Engineer",
    "Systems Engineer", "Security Engineer", "Solutions Architect",
]

# ══════════════════════════════════════════════════════════════════════
# NOISE / SPAM FILTERING
# ══════════════════════════════════════════════════════════════════════
# Sender domains that are never job-related — pre-filtered before classification
NOISE_SENDER_DOMAINS = {
    # Banking & Finance
    "hdfcbank.bank.in", "mailers.hdfcbank.bank.in", "icicibank.com", "axisbank.com",
    "sbi.co.in", "kotak.com", "yesbank.in", "indusind.com", "cdslindia.co.in",
    "groww.in", "zerodha.com", "upstox.com", "kite.zerodha.com", "angelone.in",
    # Shopping & E-commerce
    "amazon.in", "amazon.com", "flipkart.com", "rmp.flipkart.com", "myntra.com",
    "ajio.com", "comm.adidas.in", "adidas.in", "bluestone.com", "nykaa.com",
    "tatacliq.com", "reliancedigital.in", "croma.com", "meesho.com",
    # Food & Delivery
    "swiggy.in", "zomato.com", "updates.rapido.bike", "rapido.bike",
    "dunzo.com", "blinkit.com", "bigbasket.com",
    # Streaming & Social & Gaming
    "twitch.tv", "discord.com", "update.strava.com", "strava.com",
    "spotify.com", "netflix.com", "hotstar.com", "primevideo.com",
    "youtube.com", "facebook.com", "instagram.com",
    # Travel
    "makemytrip.com", "zen-makemytrip.com", "goindigo.in", "exclusive.goindigo.in",
    "cleartrip.com", "yatra.com", "irctc.co.in", "airindia.in",
    # Real Estate & Utilities
    "homeservices.nobroker.in", "nobroker.in", "magicbricks.com", "99acres.com",
    # Hardware & Gaming
    "newsletter.originpc.com", "originpc.com", "steam.com", "epicgames.com",
    # Telecom
    "jio.com", "airtel.in", "vodafone.in",
    # Security / System notifications (not job-related)
    "accounts.google.com", "accountprotection.microsoft.com",
    "no-reply@microsoft.com",
    # Consumer Electronics
    "in.email.samsung.com", "samsung.com",
    # Delivery status / Mailer daemon
    "googlemail.com",
}

# Subject-line keywords that indicate non-job emails
NOISE_SUBJECT_KEYWORDS = [
    # Shopping & Sales
    "cashback", "flat % off", "flat 40%", "flat 50%", "price drop", "voucher",
    "coupon", "discount code", "limited time offer", "flash sale", "end of season",
    "shop now", "buy now", "free delivery", "free shipping",
    # Banking & Finance
    "payment was made", "credit card xx", "emi", "loan processing",
    "smartemi", "outstanding balance", "e-voting", "nominee",
    "add nominee", "statement ready", "transaction alert",
    # Streaming & Gaming
    "is live:", "gave you kudos", "mentioned you in", "new follower",
    "started streaming", "went live",
    # E-commerce Orders
    "ordered:", "shipped:", "delivered:", "your order", "order confirmed",
    "out for delivery", "refund processed",
    # Travel
    "flight booking", "booking confirmed", "itinerary",
    # Generic Spam
    "unsubscribe", "huge savings", "exclusive offer", "act now",
    # Security / System alerts
    "security alert", "new sign-in detected", "two-step verification",
    "security code", "security info was added", "new app(s) connected",
    "run failed:", "delivery status notification",
    # Consumer marketing
    "samsung ai", "switch to samsung",
]

# Recruiter ATS / hiring platform domains — bonus scoring weight
RECRUITER_PLATFORM_DOMAINS = {
    "greenhouse.io", "lever.co", "smartrecruiters.com", "ashbyhq.com",
    "workday.com", "icims.com", "bamboohr.com", "jazz.co", "breezy.hr",
    "recruitee.com", "workable.com", "teamtailor.com", "personio.de",
    "hire.lever.co", "boards.greenhouse.io", "myworkdayjobs.com",
    "jobs.lever.co", "apply.workable.com",
}

# ══════════════════════════════════════════════════════════════════════
# MULTI-LINGUAL INTENT DICTIONARIES
# ══════════════════════════════════════════════════════════════════════
MULTILINGUAL_SIGNALS = {
    "interview": {
        "en": [
            "invitation to interview", "invite you to interview", "pleased to invite you",
            "schedule a call", "schedule an interview", "schedule a time",
            "book a time", "select a time slot", "choose a time",
            "like to speak with you", "like to chat with you", "love to connect",
            "next step in our process", "moving forward to the next round",
            "next round of interviews", "first round interview", "screening interview",
            "introductory call", "hiring manager interview", "technical interview round",
            "calendly.com", "goodtime.io", "savvycal.com", "schedule your interview",
            # New expanded English signals
            "would love to set up a call", "let's set up a conversation",
            "would like to invite you", "interview schedule", "we are pleased to inform",
            "shortlisted for interview", "selected for the next round",
            "your profile has been shortlisted", "confirm your interview",
            "proceed to the next step", "proceed to next round",
            "we'd like to schedule", "we would like to schedule",
            "let's arrange a call", "arrange a meeting", "arrange an interview",
            "discussion with the team", "meet the team", "panel interview",
            "final round interview", "on-site interview", "video interview",
            "phone screening", "phone screen", "initial call", "discovery call",
            "chat about the role", "chat about this opportunity",
            "connect with our hiring", "speak with our recruiter",
            "interview slot", "interview time", "interview date",
            "we want to move forward", "move to the next stage",
            "advance your application", "advance to the next step",
            "pleased to inform you that you have been shortlisted",
            "happy to let you know", "glad to inform you",
        ],
        "de": [
            "einladung zum interview", "einladung zum vorstellungsgespräch", "einladung zum kennenlerngespräch",
            "termin für ein interview", "telefoninterview", "erstgespräch", "kennenlernen möchten",
            "nächste runde", "terminvereinbarung", "interviewtermin", "vorstellungsgespräch vereinbaren",
            "wir möchten sie zu einem interview einladen", "einen termin abstimmen",
            "wir laden sie herzlich ein", "gesprächstermin", "bewerbungsgespräch",
        ],
        "fr": [
            "invitation à un entretien", "invitation pour un entretien", "entretien de pré-qualification",
            "entretien téléphonique", "prochaine étape de notre processus", "convenir d'un rendez-vous",
            "échangeons par téléphone", "entretien avec le responsable", "planifier un appel",
            "nous souhaitons échanger avec vous", "fixer un entretien",
            "nous avons le plaisir de vous convier", "convocation à un entretien",
            "entretien de recrutement", "planifier un entretien", "rendez-vous d'entretien",
            "interview schedule", "entretien rh",
        ],
        "nl": [
            "uitnodiging voor een gesprek", "uitnodiging voor een sollicitatiegesprek", "kennismakingsgesprek",
            "telefonisch interview", "volgende ronde", "afspraak inplannen", "gesprek plannen",
            "graag met je in gesprek", "uitnodigen voor een interview",
        ],
        "nordic": [
            "inbjudan till intervju", "boka en tid för intervju", "intervju med oss",
            "inbjudan till samtal", "nästa steg i rekryteringen", "kutsu haastatteluun",
            "haastattelukutsu", "kutsumme sinut haastatteluun", "til samtale", "innkalling til intervju",
        ],
        "ar": [
            "دعوة للمقابلة", "مقابلة شخصية", "الخطوة التالية في التوظيف",
            "تحديد موعد للمقابلة", "مقابلة مبدئية", "يسعدنا دعوتك للمقابلة",
            "لقاء تقني", "موعد مقابلة", "جدولة مكالمة",
        ],
        "es_pt_it_pl": [
            "invitación a una entrevista", "invitamos a una entrevista", "agendar una entrevista",
            "siguiente paso en el proceso", "llamada de selección", "reunión de presentación",
            "convite para entrevista", "convidamos para uma entrevista", "agendar uma conversa",
            "próxima etapa do processo", "invito a un colloquio", "fissare un colloquio",
            "zaproszenie na rozmowę", "zaproszenie do kolejnego etapu", "rozmowa rekrutacyjna",
        ]
    },
    "assessment": {
        "en": [
            "technical assessment", "coding challenge", "coding test",
            "hackerrank", "codility", "testgorilla", "byteboard",
            "take-home test", "take home assessment", "practical test",
            "online assessment", "complete the assessment before", "assessment deadline",
            "hirevue", "coderbyte", "codewars",
            # New expanded signals
            "aptitude test", "skill assessment", "pre-screening test",
            "complete the following", "complete this assessment",
            "codesignal", "leetcode", "amcat", "cocubes", "mettl",
            "technical test link", "assessment link", "test invitation",
            "online test", "programming test", "code test", "assessment invitation",
            "complete your assessment", "start your assessment",
            "technical evaluation", "skills evaluation", "coding exercise",
            "code challenge", "home assignment", "take-home assignment",
            "technical task", "screening assessment",
        ],
        "de": [
            "technischer test", "programmiertest", "coding aufgabe", "coding challenge",
            "online test", "praxisaufgabe", "fallstudie", "online-assessment",
        ],
        "fr": [
            "test technique", "évaluation technique", "test de codage", "test en ligne",
            "cas pratique", "mise en situation", "évaluation de compétences",
            "évaluation des compétences", "test de compétences",
        ],
        "nl": [
            "technische test", "codeertest", "programmeertest", "online assessment", "praktijkopdracht",
        ],
        "nordic": [
            "tekniskt test", "programmeringstest", "teknisk uppgift", "tekninen tehtävä", "koodaustesti",
        ],
        "ar": [
            "اختبار تقني", "اختبار برمجي", "تقييم فني", "تحدي برمجي", "اختبار مهارات",
        ],
        "es_pt_it_pl": [
            "prueba técnica", "test técnico", "teste técnico", "desafio técnico",
            "test tecnico", "prova tecnica", "zadanie rekrutacyjne", "test techniczny",
        ]
    },
    "inquiry": {
        "en": [
            "please let us know your availability", "share your availability",
            "provide your availability", "what is your notice period",
            "expected salary", "salary expectations", "right to work",
            "visa sponsorship required", "share your updated cv",
            "could you confirm your current location", "share a few time slots",
            # New expanded signals
            "send us your updated resume", "current ctc", "expected ctc",
            "notice period", "when can you join", "share your resume",
            "please confirm your details", "please confirm details",
            "earliest joining date", "joining date", "current salary",
            "desired salary", "salary range", "compensation expectations",
            "relocation", "willing to relocate", "open to relocation",
            "work authorization", "work permit", "sponsorship",
            "updated cv", "latest resume", "updated resume",
            "confirm your contact", "confirm your phone", "contact number",
            "please share your", "kindly share your", "kindly provide",
            "could you share", "respond with your", "reply with your availability",
        ],
        "de": [
            "verfügbarkeit", "terminvorschläge", "kündigungsfrist", "gehaltsvorstellung",
            "gehaltswunsch", "arbeitserlaubnis", "ab wann könnten sie beginnen",
        ],
        "fr": [
            "vos disponibilités", "créneaux disponibles", "période de préavis",
            "délai de préavis", "prétentions salariales", "autorisation de travail",
            "date de disponibilité", "disponible à partir de",
        ],
        "nl": [
            "beschikbaarheid", "beschikbare momenten", "opzegtermijn", "salarisindicatie", "werkvergunning",
        ],
        "nordic": [
            "tillgänglighet", "uppsägningstid", "löneanspråk", "työlupa", "irtisanomisaika",
        ],
        "ar": [
            "توفرك", "أوقات فراغك", "فترة الإشعار", "الراتب المتوقع", "تصريح العمل", "تأشيرة العمل",
        ],
        "es_pt_it_pl": [
            "disponibilidad", "período de preaviso", "expectativas salariales",
            "disponibilidade", "aviso prévio", "pretensão salarial",
            "disponibilità", "preavviso", "dostępność", "okres wypowiedzenia",
        ]
    },
    "rejection": {
        "en": [
            "unfortunately", "not moving forward", "other candidates",
            "pursue other candidates", "decided not to proceed",
            "not selected for this role", "carefully reviewed your application",
            "at this time we have decided", "wish you all the best",
            "will not be progressing", "regret to inform",
            "unable to move forward", "will not be moving forward",
            "not be proceeding", "decided to go with another", "chose another candidate",
            "no longer being considered", "position has been filled",
            "role has been filled", "won't be advancing",
            "not a fit at this time", "decided to proceed with other",
            "unable to offer you", "not advancing", "unsuccessful", "not successful",
            "application was not successful", "after careful consideration",
            "not able to offer", "chosen to move forward with other",
            "we are unable to proceed", "will not be moving forward with your candidacy",
            "we have chosen to move forward with another",
            "although your experience is impressive",
            "we've decided to pursue other", "not moving you forward",
            "not moving ahead", "we won't be moving forward",
        ],
        "de": [
            "leider müssen wir ihnen mitteilen", "haben uns für einen anderen kandidaten entschieden",
            "nicht weiterverfolgen", "bedauern wir ihnen mitteilen", "absage",
            "leider nicht berücksichtigen", "nicht in die engere auswahl", "anderweitig entschieden",
            "bewerbung leider absagen", "wir bedauern sehr", "ihre bewerbung konnte leider nicht",
            "nicht fortführen", "leider eine absage", "konnten ihre bewerbung leider nicht",
        ],
        "fr": [
            "nous avons le regret de vous informer", "nous ne retenons pas votre candidature",
            "malheureusement", "poursuivre avec d'autres candidats", "pas retenu votre profil",
            "nous ne sommes pas en mesure de donner suite", "candidature n'a pas été retenue",
            "ne donnera pas suite", "d'autres candidatures correspondaient davantage",
            "refus de candidature", "votre candidature n'a pas été sélectionnée",
            "nous sommes au regret", "ne donnera pas une suite favorable",
            "ne pouvons donner une suite favorable", "pas été retenu",
        ],
        "nl": [
            "helaas moeten wij u mededelen", "gekozen voor een andere kandidaat",
            "niet verder gaan met je sollicitatie", "afwijzing", "helaas niet geselecteerd",
            "niet in aanmerking", "we hebben besloten niet verder te gaan", "helaas afwijzen",
            "sollicitatie niet voortzetten",
        ],
        "nordic": [
            "tyvärr har vi valt att gå vidare med andra", "vi kan tyvärr inte erbjuda dig",
            "emme valitettavasti etene", "valitettavasti emme voi", "avslag på søknad",
            "ikke kommet videre", "inte gått vidare", "valitettavasti emme voi tarjota",
        ],
        "ar": [
            "للأسف نعتذر عن عدم المضي قدماً", "قررنا المضي مع مرشحين آخرين",
            "نعتذر عن عدم قبول الطلب", "نتمنى لك التوفيق في فرص أخرى",
            "نأسف لإبلاغك", "نعتذر عن المضي قدما",
        ],
        "es_pt_it_pl": [
            "lamentamos informarle", "no avanzaremos con su candidatura",
            "infelizmente não seguiremos", "purtroppo non proseguiremo",
            "niestety nie możemy przejść dalej", "nie zakwalifikował się",
            "non è stata selezionata", "candidatura non accolta", "odrzucenie aplikacji",
            "lamentamos comunicar", "não seguiremos com o processo",
        ]
    },
    "confirmation": {
        "en": [
            "thank you for applying", "application received", "we have received your application",
            "application submitted", "thanks for applying", "your application has been received",
            # New expanded signals
            "application is under review", "reviewing your application",
            "application is being reviewed", "application has been submitted",
            "successfully submitted your application", "application was sent to",
            "applied for the position", "your resume has been received",
            "acknowledge receipt", "we acknowledge your application",
        ],
        "de": [
            "vielen dank für ihre bewerbung", "bewerbungseingang", "eingangsbestätigung", "erfolgreich eingegangen",
        ],
        "fr": [
            "nous accusons réception de votre candidature", "merci pour votre candidature", "candidature bien reçue",
            "bienvenue à bord du processus de recrutement", "nous avons bien reçu votre candidature",
            "bien recu votre candidature", "candidature a été enregistrée",
            "avons bien reçu votre candidature",
        ],
        "nl": [
            "bedankt voor je sollicitatie", "sollicitatie goed ontvangen", "ontvangstbevestiging",
        ],
        "nordic": [
            "tack för din ansökan", "kiitos hakemuksestasi", "mottatt søknad",
        ],
        "ar": [
            "شكراً لتقديمك", "تم استلام طلبك بنجاح", "تأكيد استلام طلب التوظيف",
        ],
        "es_pt_it_pl": [
            "gracias por postular", "recebemos sua candidatura",
            "grazie per aver inviato la candidatura", "dziękujemy za przesłanie aplikacji",
        ]
    },
    "recruiter_outreach": {
        "en": [
            "your background could be a great match", "job invite from recruiter",
            "your profile is shortlisted", "you've been chosen", "you've been shortlisted",
            "we'd like to connect", "interested in your profile",
            "we found your profile", "your resume caught our attention",
            "we have a suitable opening", "your profile caught our eye",
            "we came across your profile", "you appear to be a strong fit",
            "excited to share this opportunity", "great match for this role",
            "ideal candidate for", "perfect fit for",
            "reach out regarding", "reaching out about an opportunity",
            "i noticed your experience", "your skills align",
            "we think you'd be a great fit", "profile is shortlisted",
            "shortlisted! please confirm", "please confirm details",
            "i am a sr. recruitment consultant", "while reviewing top talents",
            "your profile has been selected", "we have identified your profile",
        ],
        "fr": [
            "votre profil a retenu notre attention", "nous avons identifié votre profil",
            "votre candidature nous intéresse", "votre profil correspond",
            "nous avons trouvé votre profil", "votre expérience nous intéresse",
        ],
        "de": [
            "ihr profil hat unser interesse geweckt", "wir haben ihr profil gefunden",
            "sie passen hervorragend", "ihre qualifikationen sind sehr interessant",
        ],
        "nl": [
            "uw profiel past bij", "we kwamen uw profiel tegen",
        ],
        "ar": [
            "ملفك الشخصي لفت انتباهنا", "نود التواصل معك بخصوص",
        ],
        "es_pt_it_pl": [
            "su perfil nos ha interesado", "seu perfil chamou nossa atenção",
            "il suo profilo ci ha interessato", "pana profil nas zainteresował",
        ]
    },
    "job_alert": {
        "en": [
            "new jobs match your preferences", "job alert for",
            "jobs you might be interested in", "jobs we think you'll like",
            "new jobs for", "job alert", "new job matches",
            "hot job opportunities", "top employers hiring",
            "job opportunities waiting", "similar jobs in",
            "companies are looking for candidates like you",
            "jobs matching your profile",
        ],
        "fr": [
            "nouvelles offres correspondent", "alerte emploi",
            "offres d'emploi correspondant à votre profil",
        ],
        "de": [
            "neue stellen passend zu ihrem profil", "jobalarm",
            "neue jobangebote",
        ],
    }
}


def log(msg: str, level: str = "INFO"):
    """Formatted timestamped console logger."""
    ts = datetime.now().strftime("%H:%M:%S")
    icons = {
        "INFO": "[i]", "OK": "[OK]", "WARN": "[!]",
        "ERROR": "[ERR]", "NEXT": "[NEXT]", "TEST": "[TEST]"
    }
    icon = icons.get(level, "[i]")
    safe = str(msg).encode("ascii", errors="replace").decode("ascii")
    print(f"[{ts}] {icon}  {safe}", flush=True)


# ══════════════════════════════════════════════════════════════════════
# MULTI-ACCOUNT CONFIGURATION LOADER
# ══════════════════════════════════════════════════════════════════════
def resolve_server_presets(account: dict) -> dict:
    """Ensure imap_server and imap_port are configured based on provider preset."""
    provider = str(account.get("provider", "gmail")).lower().strip()
    presets = {
        "gmail": ("imap.gmail.com", 993),
        "google": ("imap.gmail.com", 993),
        "outlook": ("outlook.office365.com", 993),
        "hotmail": ("outlook.office365.com", 993),
        "office365": ("outlook.office365.com", 993),
        "microsoft": ("outlook.office365.com", 993),
        "yahoo": ("imap.mail.yahoo.com", 993),
        "icloud": ("imap.mail.me.com", 993),
    }

    if not account.get("imap_server") and provider in presets:
        account["imap_server"], account["imap_port"] = presets[provider]
    elif not account.get("imap_port"):
        account["imap_port"] = 993

    if not account.get("name"):
        account["name"] = f"{provider.capitalize()} ({account.get('email', 'account')})"

    if not account.get("folders"):
        account["folders"] = ["INBOX"]

    return account


def load_email_config() -> dict:
    """
    Load multi-account configuration supporting:
      1. `email_config.json` with `accounts: [...]` list (1 Gmail + multiple Outlook)
      2. `.env` file variables (GMAIL_*, OUTLOOK1_*, OUTLOOK2_*)
      3. Legacy single-account fallback
    """
    config = {
        "accounts": [],
        "days_back": 14,
        "max_emails_per_account": 150,
        "mark_as_read": False,
    }

    # 1. Check local email_config.json (ignored by git)
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                disk_cfg = json.load(f)

                if "accounts" in disk_cfg and isinstance(disk_cfg["accounts"], list):
                    for acc in disk_cfg["accounts"]:
                        if acc.get("enabled", True) and (acc.get("email") or acc.get("password")):
                            config["accounts"].append(resolve_server_presets(acc))
                elif disk_cfg.get("email") and disk_cfg.get("password"):
                    # Single account legacy format
                    config["accounts"].append(resolve_server_presets(disk_cfg))

                for k in ["days_back", "max_emails_per_account", "mark_as_read"]:
                    if k in disk_cfg:
                        config[k] = disk_cfg[k]
        except Exception as e:
            log(f"Warning reading {CONFIG_PATH}: {e}", "WARN")

    # 2. Check .env for multi-account environment variables if accounts empty
    if not config["accounts"] and ENV_PATH.exists():
        try:
            env_vars = {}
            with open(ENV_PATH, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    k, v = line.split("=", 1)
                    env_vars[k.strip().upper()] = v.strip().strip("\"'")

            # Account 1: Gmail
            if env_vars.get("GMAIL_USER") and env_vars.get("GMAIL_PASSWORD"):
                config["accounts"].append(resolve_server_presets({
                    "name": "Gmail Primary",
                    "email": env_vars["GMAIL_USER"],
                    "password": env_vars["GMAIL_PASSWORD"],
                    "provider": "gmail"
                }))

            # Account 2: Outlook 1
            if env_vars.get("OUTLOOK1_USER") and env_vars.get("OUTLOOK1_PASSWORD"):
                config["accounts"].append(resolve_server_presets({
                    "name": "Outlook Account 1",
                    "email": env_vars["OUTLOOK1_USER"],
                    "password": env_vars["OUTLOOK1_PASSWORD"],
                    "provider": "outlook"
                }))

            # Account 3: Outlook 2
            if env_vars.get("OUTLOOK2_USER") and env_vars.get("OUTLOOK2_PASSWORD"):
                config["accounts"].append(resolve_server_presets({
                    "name": "Outlook Account 2",
                    "email": env_vars["OUTLOOK2_USER"],
                    "password": env_vars["OUTLOOK2_PASSWORD"],
                    "provider": "outlook"
                }))

            # General EMAIL_USER fallback
            if not config["accounts"] and env_vars.get("EMAIL_USER") and env_vars.get("EMAIL_PASSWORD"):
                config["accounts"].append(resolve_server_presets({
                    "name": "Primary Account",
                    "email": env_vars["EMAIL_USER"],
                    "password": env_vars["EMAIL_PASSWORD"],
                    "provider": env_vars.get("EMAIL_PROVIDER", "gmail")
                }))
        except Exception as e:
            log(f"Warning reading .env: {e}", "WARN")

    return config


# ══════════════════════════════════════════════════════════════════════
# EMAIL DECODING & PARSING
# ══════════════════════════════════════════════════════════════════════
def decode_mime_header(header_val: str) -> str:
    """Decode RFC 2047 MIME encoded headers safely."""
    if not header_val:
        return ""
    decoded_parts = []
    try:
        parts = decode_header(header_val)
        for text, charset in parts:
            if isinstance(text, bytes):
                try:
                    charset = charset or "utf-8"
                    decoded_parts.append(text.decode(charset, errors="replace"))
                except Exception:
                    decoded_parts.append(text.decode("latin-1", errors="replace"))
            else:
                decoded_parts.append(str(text))
    except Exception:
        return str(header_val)
    return " ".join(decoded_parts).strip()


def extract_email_body(msg: email.message.Message) -> tuple[str, str]:
    """Extract plain text and HTML bodies from a multi-part email message."""
    plain_text = ""
    html_text = ""

    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition", ""))

            if "attachment" in content_disposition:
                continue

            payload = part.get_payload(decode=True)
            if not payload:
                continue

            charset = part.get_content_charset() or "utf-8"
            try:
                decoded = payload.decode(charset, errors="replace")
            except Exception:
                decoded = payload.decode("latin-1", errors="replace")

            if content_type == "text/plain" and not plain_text:
                plain_text = decoded
            elif content_type == "text/html" and not html_text:
                html_text = decoded
    else:
        content_type = msg.get_content_type()
        payload = msg.get_payload(decode=True)
        if payload:
            charset = msg.get_content_charset() or "utf-8"
            try:
                decoded = payload.decode(charset, errors="replace")
            except Exception:
                decoded = payload.decode("latin-1", errors="replace")

            if content_type == "text/html":
                html_text = decoded
            else:
                plain_text = decoded

    # Clean HTML to plain text fallback if text/plain is missing
    if not plain_text and html_text:
        clean = re.sub(r'<style.*?>.*?</style>', '', html_text, flags=re.DOTALL | re.IGNORECASE)
        clean = re.sub(r'<script.*?>.*?</script>', '', clean, flags=re.DOTALL | re.IGNORECASE)
        clean = re.sub(r'<br\s*/?>', '\n', clean, flags=re.IGNORECASE)
        clean = re.sub(r'</p>', '\n\n', clean, flags=re.IGNORECASE)
        clean = re.sub(r'<[^>]+>', ' ', clean)
        clean = re.sub(r'&nbsp;', ' ', clean)
        clean = re.sub(r'&amp;', '&', clean)
        clean = re.sub(r'&lt;', '<', clean)
        clean = re.sub(r'&gt;', '>', clean)
        clean = re.sub(r'\n{3,}', '\n\n', clean)
        plain_text = clean.strip()

    return plain_text.strip(), html_text.strip()


def extract_action_links(plain_text: str, html_text: str) -> list[str]:
    """Extract scheduling, interview, and assessment links from email."""
    combined = f"{plain_text}\n{html_text}"
    links = []

    for pattern in MEETING_LINK_PATTERNS:
        matches = re.findall(pattern, combined, flags=re.IGNORECASE)
        for m in matches:
            clean_link = m.rstrip('.,;)"\'\\]>')
            if clean_link not in links:
                links.append(clean_link)

    # Search for calendar/interview href keywords in HTML
    if html_text:
        href_matches = re.findall(r'href=[\'"](https?://[^\'"]+)[\'"]', html_text, flags=re.IGNORECASE)
        for href in href_matches:
            if any(k in href.lower() for k in ["schedule", "calendly", "goodtime", "interview", "assessment", "booking", "meet.google", "teams.microsoft", "zoom.us"]):
                if href not in links and not any(href.endswith(ext) for ext in [".png", ".jpg", ".jpeg", ".gif", ".css", ".js"]):
                    links.append(href)

    return links


# ══════════════════════════════════════════════════════════════════════
# NOISE FILTERING & MULTI-LINGUAL CLASSIFICATION ENGINE
# ══════════════════════════════════════════════════════════════════════
def is_noise_email(from_header: str, subject: str) -> bool:
    """Pre-filter: returns True if the email is clearly non-job-related noise."""
    # Check sender domain against noise blocklist
    email_match = re.search(r'@([a-zA-Z0-9.-]+)', from_header)
    if email_match:
        sender_domain = email_match.group(1).lower()
        for noise_domain in NOISE_SENDER_DOMAINS:
            if sender_domain == noise_domain or sender_domain.endswith("." + noise_domain):
                return True

    # Check subject against noise keywords
    subj_lower = subject.lower()
    for keyword in NOISE_SUBJECT_KEYWORDS:
        if keyword in subj_lower:
            return True

    return False


def detect_language(text: str) -> str:
    """Detect language based on multi-lingual keywords."""
    text_lower = text.lower()
    if bool(re.search(r'[\u0600-\u06FF]', text)):
        return "Arabic"
    if any(k in text_lower for k in ["bewerbung", "vorstellungsgespräch", "herzlichen dank", "kündigungsfrist", "gehaltsvorstellung", "einladung", "lebenslauf"]):
        return "German"
    if any(k in text_lower for k in ["candidature", "entretien", "disponibilités", "prétentions salariales", "cordialement", "recrutement", "poste de", "ingénieur", "bienvenue"]):
        return "French"
    if any(k in text_lower for k in ["sollicitatie", "kennismakingsgesprek", "beschikbaarheid", "opzegtermijn", "met vriendelijke groet"]):
        return "Dutch"
    if any(k in text_lower for k in ["ansökan", "intervju", "tillgänglighet", "hakemuksesi", "haastattelukutsu"]):
        return "Nordic"
    if any(k in text_lower for k in ["entrevista", "candidatura", "disponibilidad", "preaviso", "atentamente"]):
        return "Spanish / Portuguese"
    if any(k in text_lower for k in ["colloquio", "disponibilità", "cordiali saluti"]):
        return "Italian"
    if any(k in text_lower for k in ["rekrutacja", "rozmowa kwalifikacyjna", "dostępność", "pozdrawiamy"]):
        return "Polish"
    return "English"


def extract_company_name(from_header: str, subject: str, body: str) -> str:
    """Extract hiring company name from headers, sender domain, or subject."""
    subject_patterns = [
        r'(?:interview|chat|invitation|speaking|gespräch|entretien|entrevista)\s+(?:with|at|bei|avec|con)\s+([A-Z0-9][A-Za-z0-9\s&.,-]{2,30}?)(?:\s+for|\s+-|\s+:|\s+team|\s+für|\s+pour|\s*$)',
        r'(?:your\s+application\s+(?:to|at|with)|ihre\s+bewerbung\s+bei|votre\s+candidature\s+(?:chez|auprès\s+de))\s+([A-Z0-9][A-Za-z0-9\s&.,-]{2,30}?)(?:\s+for|\s+-|\s+:|\s+als|\s*$)',
        r'\[([A-Za-z0-9\s&.,-]{2,25})\]\s+(?:interview|application|bewerbung|candidature|next\s+step)',
    ]
    for pat in subject_patterns:
        m = re.search(pat, subject, flags=re.IGNORECASE)
        if m:
            c = m.group(1).strip(" -:;,.")
            if len(c) >= 2 and not any(w in c.lower() for w in ["interview", "application", "bewerbung", "candidature", "update", "reminder"]):
                return c

    # From Display Name: "John Doe from Stripe", "Google Careers"
    if "<" in from_header:
        display_name = from_header.split("<")[0].strip("\"' ")
        email_addr = from_header.split("<")[1].rstrip(">").strip()
    else:
        display_name = from_header
        email_addr = from_header

    m_from = re.search(r'(?:from|at|@|bei|de|chez)\s+([A-Z0-9][A-Za-z0-9\s&.,-]{2,30})', display_name, flags=re.IGNORECASE)
    if m_from:
        return m_from.group(1).strip(" -:;,.")

    # From Email Domain
    domain_match = re.search(r'@([a-zA-Z0-9.-]+)\.([a-zA-Z]{2,})', email_addr)
    if domain_match:
        domain_name = domain_match.group(1).lower()
        generic_domains = ["gmail", "yahoo", "hotmail", "outlook", "live", "icloud", "mail", "greenhouse", "lever", "workday", "smartrecruiters", "ashbyhq"]
        if domain_name not in generic_domains and len(domain_name) >= 3:
            return domain_name.capitalize()

    return "Recruiter / Hiring Team"


def extract_job_title(subject: str, body: str) -> str:
    """Identify the target job title from subject or body snippet."""
    combined = f"{subject}\n{body[:500]}"
    for role in KNOWN_TECH_ROLES:
        if re.search(rf'\b{re.escape(role)}\b', combined, flags=re.IGNORECASE):
            return role

    m = re.search(r'(?:for\s+(?:the\s+)?(?:position\s+of\s+|role\s+of\s+)?|als\s+|pour\s+le\s+poste\s+de\s+)([A-Z][A-Za-z0-9\s/&-]{3,35}?)(?:\s+role|\s+position|\s+at|\s+with|\s+-|\s*$|\n)', subject)
    if m:
        t = m.group(1).strip(" -:;,.")
        if len(t) >= 4 and not any(w in t.lower() for w in ["your", "next", "update", "application", "ihre", "votre", "schedule"]):
            return t

    return "DevOps / Cloud Role"


def classify_email(subject: str, body: str, links: list[str], from_header: str = "") -> dict:
    """Multi-lingual classification with weighted scoring, HTML entity decoding, and noise filtering."""

    # Step 1: Decode HTML entities so French accents (é, ç, à) match signal phrases
    raw_combined = f"{subject}\n{body}"
    decoded_combined = html_module.unescape(raw_combined)
    text = decoded_combined.lower()
    subject_lower = html_module.unescape(subject).lower()

    lang = detect_language(decoded_combined)

    # Step 2: Weighted scoring — subject matches get 3x, body matches get 2x
    def score_category(cat_name):
        cat_dict = MULTILINGUAL_SIGNALS.get(cat_name, {})
        subject_score = 0
        body_score = 0
        for _, phrases in cat_dict.items():
            for p in phrases:
                p_lower = p.lower()
                if p_lower in subject_lower:
                    subject_score += 3  # Subject match = strong signal
                elif p_lower in text:
                    body_score += 2     # Body match = moderate signal
        return subject_score + body_score

    # Step 3: Compute scores for all categories
    interview_score = score_category("interview")
    assessment_score = score_category("assessment")
    inquiry_score = score_category("inquiry")
    rejection_score = score_category("rejection")
    confirmation_score = score_category("confirmation")
    outreach_score = score_category("recruiter_outreach")
    job_alert_score = score_category("job_alert")

    # Step 4: Link presence bonuses
    meeting_link_keywords = ["calendly", "goodtime", "savvycal", "chilipiper", "schedule",
                              "meet.google", "teams.microsoft", "zoom.us", "webex"]
    assessment_link_keywords = ["hackerrank", "codility", "testgorilla", "byteboard",
                                 "hirevue", "codesignal", "coderbyte", "mettl", "amcat"]

    if any(lnk for lnk in links if any(k in lnk.lower() for k in meeting_link_keywords)):
        interview_score += 5

    if any(lnk for lnk in links if any(k in lnk.lower() for k in assessment_link_keywords)):
        assessment_score += 5

    # Step 5: Recruiter platform domain bonus (sender is from ATS)
    if from_header:
        sender_domain_match = re.search(r'@([a-zA-Z0-9.-]+)', from_header)
        if sender_domain_match:
            sender_domain = sender_domain_match.group(1).lower()
            for ats_domain in RECRUITER_PLATFORM_DOMAINS:
                if sender_domain == ats_domain or sender_domain.endswith("." + ats_domain):
                    # ATS senders boost interview/assessment/confirmation scores
                    interview_score += 2
                    assessment_score += 1
                    confirmation_score += 1
                    break

    # Step 6: Classification decision tree (lowered thresholds)
    # Rejection: captures single-phrase multi-lingual rejections
    if rejection_score >= 2 and interview_score < 2 and assessment_score < 2:
        return {
            "category": "Rejection",
            "priority": "🔴 Low - Archived",
            "action_required": "Application not moving forward",
            "is_next_step": False,
            "badge": "🔴 Rejection",
            "language": lang
        }

    # Interview: lowered threshold — single strong subject match is enough
    if interview_score >= 2 or (links and interview_score >= 1):
        action = "Book interview slot via scheduling link" if links else "Reply with your availability to schedule interview"
        return {
            "category": "Interview Invitation",
            "priority": "🟢 High - Immediate Action",
            "action_required": action,
            "is_next_step": True,
            "badge": "🟢 Interview Invite",
            "language": lang
        }

    # Assessment: lowered threshold
    if assessment_score >= 2:
        return {
            "category": "Technical Assessment",
            "priority": "🔵 High - Assessment",
            "action_required": "Complete online technical assessment before deadline",
            "is_next_step": True,
            "badge": "🔵 Assessment",
            "language": lang
        }

    # Inquiry/Availability: lowered threshold
    if inquiry_score >= 2:
        return {
            "category": "Availability / Inquiry",
            "priority": "🟡 Action Required",
            "action_required": "Reply to recruiter with requested availability / details",
            "is_next_step": True,
            "badge": "🟡 Availability Request",
            "language": lang
        }

    # Recruiter Outreach: personal recruiter interest (is_next_step = True)
    if outreach_score >= 2:
        return {
            "category": "Recruiter Outreach",
            "priority": "🟠 Recruiter Interest",
            "action_required": "Review and respond to recruiter — potential opportunity",
            "is_next_step": True,
            "badge": "🟠 Recruiter Outreach",
            "language": lang
        }

    # Application Confirmation: lowered threshold
    if confirmation_score >= 1:
        return {
            "category": "Application Confirmation",
            "priority": "⚪ Info Only",
            "action_required": "Application acknowledged (pending review)",
            "is_next_step": False,
            "badge": "⚪ Confirmation",
            "language": lang
        }

    # Job Alert / Newsletter
    if job_alert_score >= 2:
        return {
            "category": "Job Alert",
            "priority": "📋 Job Alert",
            "action_required": "Browse job alert for relevant openings",
            "is_next_step": False,
            "badge": "📋 Job Alert",
            "language": lang
        }

    return {
        "category": "General Update",
        "priority": "⚪ General",
        "action_required": "Review email",
        "is_next_step": False,
        "badge": "⚪ General",
        "language": lang
    }


# ══════════════════════════════════════════════════════════════════════
# MULTI-ACCOUNT INBOX SCANNER
# ══════════════════════════════════════════════════════════════════════
def scan_single_account(account: dict, days_back: int, max_emails: int) -> list[dict]:
    """Connect to a single IMAP inbox (Gmail, Outlook 1, Outlook 2) and extract triaged emails."""
    acc_name = account.get("name", "Account")
    email_user = account.get("email", "").strip()
    email_pass = account.get("password", "").strip().replace(" ", "")
    imap_server = account.get("imap_server", "imap.gmail.com").strip()
    imap_port = int(account.get("imap_port", 993))
    folders = account.get("folders", ["INBOX"])

    if not email_user or not email_pass:
        log(f"Skipping {acc_name}: Email credentials not provided.", "WARN")
        return []

    log(f"[{acc_name}] Connecting to {imap_server}:{imap_port} ({email_user}) …", "INFO")
    results = []

    try:
        mail = imaplib.IMAP4_SSL(imap_server, imap_port)
        provider = str(account.get("provider", "")).lower()
        is_microsoft = "outlook" in provider or "hotmail" in provider or "office365" in imap_server or "outlook" in imap_server

        logged_in = False

        # Attempt 1: App Password / Basic Auth
        if email_pass:
            try:
                mail.login(email_user, email_pass)
                logged_in = True
                log(f"[{acc_name}] IMAP Login Successful!", "OK")
            except Exception as pass_err:
                log(f"[{acc_name}] App Password login failed: {pass_err}", "WARN")

        # Attempt 2: Microsoft Modern Auth (XOAUTH2)
        if not logged_in and is_microsoft:
            try:
                from microsoft_auth import get_microsoft_access_token
                token = get_microsoft_access_token(email_user, interactive=False)
                if token:
                    auth_str = f"user={email_user}\x01auth=Bearer {token}\x01\x01".encode("utf-8")
                    mail.authenticate("XOAUTH2", lambda _: auth_str)
                    logged_in = True
                    log(f"[{acc_name}] Microsoft Modern Auth (XOAUTH2) Successful!", "OK")
            except Exception as oauth_err:
                log(f"[{acc_name}] XOAUTH2 Failed: {oauth_err}", "WARN")

        if not logged_in:
            log(f"[{acc_name}] Could not authenticate with {email_user}.", "ERROR")
            return []
    except Exception as conn_err:
        log(f"[{acc_name}] Connection error to {imap_server}: {conn_err}", "ERROR")
        return []

    since_date = (datetime.now() - timedelta(days=days_back)).strftime("%d-%b-%Y")

    # Auto-discover Spam/Junk folders if available on server
    scan_folders = list(folders)
    try:
        typ, raw_folders = mail.list()
        if typ == "OK" and raw_folders:
            for rf in raw_folders:
                rf_str = rf.decode("utf-8", errors="ignore")
                if r"\Junk" in rf_str or r"\Spam" in rf_str or "spam" in rf_str.lower() or "junk" in rf_str.lower():
                    m = re.search(r'"([^"]+)"\s*$', rf_str) or re.search(r'\s+([^\s]+)\s*$', rf_str)
                    if m:
                        detected_spam_folder = m.group(1).strip('"')
                        if detected_spam_folder not in scan_folders and not detected_spam_folder.startswith("Sync Issues"):
                            scan_folders.append(detected_spam_folder)
    except Exception:
        pass

    for folder in scan_folders:
        try:
            status, _ = mail.select(f'"{folder}"', readonly=True)
            if status != "OK":
                status, _ = mail.select(folder, readonly=True)
            if status != "OK":
                log(f"[{acc_name}] Could not open folder '{folder}'", "WARN")
                continue

            search_query = f'(SINCE "{since_date}")'
            typ, data = mail.search(None, search_query)

            if typ != "OK" or not data[0]:
                log(f"[{acc_name}] No emails in '{folder}' since {since_date}.", "INFO")
                continue

            mail_ids = data[0].split()
            log(f"[{acc_name}] Found {len(mail_ids)} total emails in {folder}. Triaging latest {min(len(mail_ids), max_emails)} …", "INFO")

            for mid in reversed(mail_ids[-max_emails:]):
                try:
                    res, msg_data = mail.fetch(mid, "(RFC822)")
                    if res != "OK" or not msg_data or not msg_data[0]:
                        continue

                    raw_email = msg_data[0][1]
                    msg = email.message_from_bytes(raw_email)

                    subject = decode_mime_header(msg.get("Subject", ""))
                    from_header = decode_mime_header(msg.get("From", ""))
                    date_header = msg.get("Date", "")
                    message_id = msg.get("Message-ID", f"{acc_name}-{mid.decode('utf-8', errors='ignore')}")

                    try:
                        date_parsed = email.utils.parsedate_to_datetime(date_header)
                        date_str = date_parsed.strftime("%Y-%m-%d %H:%M")
                        date_sort = date_parsed.strftime("%Y-%m-%d")
                    except Exception:
                        date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                        date_sort = datetime.now().strftime("%Y-%m-%d")

                    plain_body, html_body = extract_email_body(msg)

                    # Pre-filter: skip non-job noise (banking, shopping, streaming, etc.)
                    if is_noise_email(from_header, subject):
                        continue

                    links = extract_action_links(plain_body, html_body)

                    classification = classify_email(subject, plain_body, links, from_header)
                    company = extract_company_name(from_header, subject, plain_body)
                    role = extract_job_title(subject, plain_body)
                    primary_link = links[0] if links else ""
                    snippet = plain_body[:300].replace('\n', ' ').strip()

                    record = {
                        "Account": acc_name,
                        "Inbox Email": email_user,
                        "Message ID": message_id,
                        "Date Received": date_str,
                        "Date Sort": date_sort,
                        "Company": company,
                        "Job Title": role,
                        "Category": classification["category"],
                        "Priority": classification["priority"],
                        "Badge": classification["badge"],
                        "Language": classification["language"],
                        "Action Required": classification["action_required"],
                        "Is Next Step": classification["is_next_step"],
                        "Action URL": primary_link,
                        "All Links": " | ".join(links),
                        "Sender": from_header,
                        "Subject": subject,
                        "Email Snippet": snippet,
                        "Full Email Body": plain_body.strip(),
                        "Folder": folder,
                        "Replied Status": "No",
                        "Notes": "",
                    }

                    results.append(record)

                    if classification["is_next_step"]:
                        log(f"   [{acc_name}] {classification['badge']} ({classification['language']}) | {company} — '{role}'", "OK")
                        if primary_link:
                            log(f"      🔗 Link: {primary_link[:65]}...", "NEXT")

                except Exception:
                    continue

        except Exception as folder_err:
            log(f"[{acc_name}] Error scanning {folder}: {folder_err}", "WARN")

    try:
        mail.close()
        mail.logout()
    except Exception:
        pass

    log(f"[{acc_name}] Scanned {len(results)} emails successfully.\n", "OK")
    return results


def scan_all_accounts(config: dict) -> list[dict]:
    """Scan all configured accounts (1 Gmail + multiple Outlook) sequentially."""
    accounts = config.get("accounts", [])
    days_back = int(config.get("days_back", 14))
    max_emails = int(config.get("max_emails_per_account", 150))

    if not accounts:
        log("No accounts configured. Please configure your Gmail and Outlook accounts in email_config.json or .env", "WARN")
        return []

    log(f"Starting Multi-Account Scan across {len(accounts)} configured inboxes …\n", "INFO")
    all_results = []

    for acc in accounts:
        acc_results = scan_single_account(acc, days_back, max_emails)
        all_results.extend(acc_results)

    return all_results


# ══════════════════════════════════════════════════════════════════════
# MASTER SPREADSHEET & DASHBOARD OUTPUTS
# ══════════════════════════════════════════════════════════════════════
def deduplicate_pipeline(new_records: list[dict]) -> pd.DataFrame:
    """Merge newly scanned records with existing pipeline and deduplicate."""
    if not new_records and not MASTER_CSV_PATH.exists():
        return pd.DataFrame()

    incoming_df = pd.DataFrame(new_records)

    if MASTER_CSV_PATH.exists():
        try:
            master_df = pd.read_csv(MASTER_CSV_PATH, on_bad_lines="skip")
            combined_df = pd.concat([master_df, incoming_df], ignore_index=True)
        except Exception:
            combined_df = incoming_df
    else:
        combined_df = incoming_df

    if combined_df.empty:
        return combined_df

    combined_df["_dedup"] = combined_df["Message ID"].fillna("")
    mask_empty = combined_df["_dedup"] == ""
    combined_df.loc[mask_empty, "_dedup"] = (
        combined_df.loc[mask_empty, "Company"].astype(str) + "|" +
        combined_df.loc[mask_empty, "Subject"].astype(str) + "|" +
        combined_df.loc[mask_empty, "Date Sort"].astype(str)
    )

    clean_df = combined_df.drop_duplicates(subset=["_dedup"], keep="first").copy()
    clean_df.drop(columns=["_dedup"], errors="ignore", inplace=True)

    if "Date Received" in clean_df.columns:
        clean_df.sort_values(by="Date Received", ascending=False, inplace=True)

    return clean_df


def save_master_excel(df: pd.DataFrame):
    """Save clean multi-sheet Excel with Action Required + Per-Category Tabs."""
    if not HAS_EXCEL or df is None or df.empty:
        return

    log("Generating Master Multi-Account Excel: interview_pipeline.xlsx …", "INFO")
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    cols = [
        "#", "Account", "Priority", "Language", "Category", "Company", "Job Title",
        "Action Required", "Action URL", "Date Received",
        "Sender", "Subject", "Email Snippet", "Replied Status", "Notes"
    ]

    hf = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hfnt = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    blue_fill  = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    def populate_sheet(ws, sub_df):
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill = hf; c.font = hfnt
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin

        for ri, (_, row) in enumerate(sub_df.iterrows(), 2):
            vals = [
                ri - 1,
                str(row.get("Account", "")),
                str(row.get("Priority", "")),
                str(row.get("Language", "")),
                str(row.get("Category", "")),
                str(row.get("Company", "")),
                str(row.get("Job Title", "")),
                str(row.get("Action Required", "")),
                str(row.get("Action URL", "")),
                str(row.get("Date Received", "")),
                str(row.get("Sender", "")),
                str(row.get("Subject", "")),
                str(row.get("Email Snippet", "")),
                str(row.get("Replied Status", "No")),
                str(row.get("Notes", ""))
            ]
            cat = str(row.get("Category", ""))
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Calibri", size=9)
                c.border = thin
                if cat == "Interview Invitation": c.fill = green_fill
                elif cat == "Technical Assessment": c.fill = blue_fill
                elif "Availability" in cat: c.fill = yellow_fill

        widths = [4, 18, 18, 14, 18, 22, 26, 35, 40, 18, 25, 35, 45, 14, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        if len(sub_df) > 0: ws.auto_filter.ref = ws.dimensions

    # 1. Action Required (Next Steps)
    next_steps_df = df[df["Is Next Step"] == True] if "Is Next Step" in df.columns else df
    ws_action = wb.create_sheet(title="Action Required (Next Steps)")
    populate_sheet(ws_action, next_steps_df)

    # 2. All Inbound Responses
    ws_all = wb.create_sheet(title="All Responses")
    populate_sheet(ws_all, df)

    # 3. Invites
    if "Category" in df.columns:
        invites_df = df[df["Category"] == "Interview Invitation"]
        if not invites_df.empty:
            ws_inv = wb.create_sheet(title="Interview Invites")
            populate_sheet(ws_inv, invites_df)

        # 4. Assessments
        assess_df = df[df["Category"] == "Technical Assessment"]
        if not assess_df.empty:
            ws_ass = wb.create_sheet(title="Assessments")
            populate_sheet(ws_ass, assess_df)

    try:
        wb.save(MASTER_EXCEL_PATH)
        log(f"Master Excel saved -> {MASTER_EXCEL_PATH}", "OK")
    except PermissionError:
        alt = WORK_DIR / f"interview_pipeline_{int(time.time())}.xlsx"
        wb.save(alt)
        log(f"Excel file locked -- saved to {alt}", "WARN")


def save_outputs(clean_df: pd.DataFrame):
    """Save CSV, Excel, and JSON outputs for web dashboard."""
    if clean_df.empty:
        return

    clean_df.to_csv(MASTER_CSV_PATH, index=False)
    log(f"Master CSV updated  -> {MASTER_CSV_PATH} ({len(clean_df)} total triaged emails)", "OK")

    save_master_excel(clean_df)

    records = clean_df.to_dict(orient="records")
    with open(MASTER_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)
    log(f"Dashboard JSON saved -> {MASTER_JSON_PATH}", "OK")


# ══════════════════════════════════════════════════════════════════════
# CLI ENTRY POINT
# ══════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="Multi-Account & Multi-Lingual Email Triage Pipeline")
    parser.add_argument("--days", type=int, default=14, help="Number of past days to scan (default: 14)")
    parser.add_argument("--max", type=int, default=150, help="Max emails per account (default: 150)")
    parser.add_argument("--account", type=str, default="", help="Scan a specific account name only")
    args = parser.parse_args()

    cfg = load_email_config()
    if args.days: cfg["days_back"] = args.days
    if args.max: cfg["max_emails_per_account"] = args.max

    if args.account:
        cfg["accounts"] = [a for a in cfg.get("accounts", []) if a.get("name", "").lower() == args.account.lower()]

    log("=" * 70)
    log("  MULTI-ACCOUNT & MULTI-LINGUAL EMAIL TRIAGE PIPELINE")
    log(f"   Configured Inboxes: {len(cfg.get('accounts', []))} accounts (Gmail + Outlook)")
    for i, a in enumerate(cfg.get("accounts", []), 1):
        log(f"     [{i}] {a.get('name', 'Account')} — {a.get('email', '')} ({a.get('provider', '').upper()})")
    log(f"   Time Window       : Past {cfg.get('days_back', 14)} Days")
    log(f"   Supported Lngs    : English, German, French, Dutch, Nordics, Arabic, Spanish, etc.")
    log("=" * 70)

    raw_records = scan_all_accounts(cfg)
    clean_df = deduplicate_pipeline(raw_records)

    if not clean_df.empty:
        save_outputs(clean_df)

        next_steps = clean_df[clean_df["Is Next Step"] == True] if "Is Next Step" in clean_df.columns else clean_df
        invites = clean_df[clean_df["Category"] == "Interview Invitation"]
        assessments = clean_df[clean_df["Category"] == "Technical Assessment"]
        inquiries = clean_df[clean_df["Category"] == "Availability / Inquiry"]
        outreach = clean_df[clean_df["Category"] == "Recruiter Outreach"] if "Category" in clean_df.columns else pd.DataFrame()
        confirmations = clean_df[clean_df["Category"] == "Application Confirmation"] if "Category" in clean_df.columns else pd.DataFrame()
        job_alerts = clean_df[clean_df["Category"] == "Job Alert"] if "Category" in clean_df.columns else pd.DataFrame()
        rejections = clean_df[clean_df["Category"] == "Rejection"] if "Category" in clean_df.columns else pd.DataFrame()
        general = clean_df[clean_df["Category"] == "General Update"] if "Category" in clean_df.columns else pd.DataFrame()

        log("\n" + "=" * 70)
        log("  EMAIL TRIAGE COMPLETE — FULL CLASSIFICATION SUMMARY", "OK")
        log(f"   🟢 Interview Invitations  : {len(invites)}")
        log(f"   🔵 Technical Assessments  : {len(assessments)}")
        log(f"   🟡 Availability Requests  : {len(inquiries)}")
        log(f"   🟠 Recruiter Outreach     : {len(outreach)}")
        log(f"   ⚪ Application Confirms   : {len(confirmations)}")
        log(f"   📋 Job Alerts / Newsletters: {len(job_alerts)}")
        log(f"   🔴 Rejections             : {len(rejections)}")
        log(f"   ⚪ General Updates         : {len(general)}")
        log(f"   ⭐ Total Next Steps Ready : {len(next_steps)}")
        log(f"   📊 Master Excel Workbook  : {MASTER_EXCEL_PATH}")
        log(f"   📄 Master CSV Database    : {MASTER_CSV_PATH}")
        log("=" * 70)
    else:
        log("No email records to save. Ensure your 3 account credentials are configured in email_config.json", "WARN")


if __name__ == "__main__":
    main()
