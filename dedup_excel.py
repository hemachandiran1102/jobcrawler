"""
dedup_excel.py — Remove Duplicates in Excel & CSV based on LinkedIn URL
========================================================================
Scans Excel (.xlsx) and CSV files for duplicate job entries based on
normalized LinkedIn job URLs, merges user metadata (Applied Status, Notes,
earliest Crawl Date), and writes clean, duplicate-free workbooks.

Usage:
  python dedup_excel.py                                # cleans full_crawl_jobs.xlsx & full_crawl_jobs.csv
  python dedup_excel.py --file "my_jobs.xlsx"          # cleans specific file
  python dedup_excel.py --dry-run                      # scan and preview duplicates without saving
"""

import os
import sys
import re
import time
import argparse
import pandas as pd
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

WORK_DIR = Path(__file__).parent.resolve()
MASTER_EXCEL_PATH = WORK_DIR / "full_crawl_jobs.xlsx"
MASTER_CSV_PATH = WORK_DIR / "full_crawl_jobs.csv"


def log(msg, level="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    icon = {"INFO": "[i]", "OK": "[OK]", "WARN": "[!]", "ERROR": "[ERR]"}.get(level, "")
    safe = str(msg).encode("ascii", errors="replace").decode("ascii")
    print(f"[{ts}] {icon}  {safe}", flush=True)


def normalize_job_url(url: str) -> str:
    """
    Extract canonical Job URL across LinkedIn, Indeed, Glassdoor, and other boards:
      - LinkedIn:  https://www.linkedin.com/jobs/view/<JOB_ID>
      - Indeed:    https://www.indeed.com/viewjob?jk=<JK_ID>
      - Glassdoor: https://www.glassdoor.com/job-listing/?jl=<JL_ID>
    """
    u = str(url or "").strip()
    if not u or u.lower() in {"n/a", "none", "nan", "", "-"}:
        return ""

    # Indeed: jk query parameter
    m_indeed = re.search(r"[?&]jk=([a-zA-Z0-9]+)", u)
    if m_indeed:
        return f"https://www.indeed.com/viewjob?jk={m_indeed.group(1)}"

    # Glassdoor: jl parameter
    m_gd = re.search(r"(?:jl=|jobListingId=|job-listing/.*?jl=)(\d+)", u)
    if m_gd:
        return f"https://www.glassdoor.com/job-listing/?jl={m_gd.group(1)}"

    # LinkedIn: numeric job ID
    m = re.search(r"/jobs/view/(?:[^\s/?#]*-)?(\d{6,14})", u)
    if m:
        return f"https://www.linkedin.com/jobs/view/{m.group(1)}"
    m_param = re.search(r"[?&]currentJobId=(\d{6,14})", u)
    if m_param:
        return f"https://www.linkedin.com/jobs/view/{m_param.group(1)}"

    clean = u.split("#")[0]
    base = clean.split("?")[0].rstrip("/")
    if base.startswith("http://"):
        base = "https://" + base[7:]
    return base


def deduplicate_records(records: list) -> tuple:
    """
    Deduplicate a list of job dicts based on normalized Job URL.
    Returns (deduped_records, duplicates_removed_count).
    """
    unique_map = {}
    dupes_count = 0

    for rec in records:
        raw_url = str(rec.get("Job URL", rec.get("Job Link", rec.get("Link", "")))).strip()
        norm_url = normalize_job_url(raw_url)

        if norm_url:
            key = f"URL:{norm_url}"
            rec["Job URL"] = norm_url
        else:
            comp = str(rec.get("Company", "")).strip().lower()
            title = str(rec.get("Job Title", rec.get("Title", ""))).strip().lower()
            loc = str(rec.get("Location", rec.get("Country", ""))).strip().lower()
            if comp and title:
                key = f"TITLE:{comp}|{title}|{loc}"
            else:
                # Keep unmatched item
                key = f"RAW:{len(unique_map)}"

        if key in unique_map:
            dupes_count += 1
            existing = unique_map[key]

            # 1. Preserve Applied Status = Yes
            rec_applied = str(rec.get("Applied Status", rec.get("Applied", ""))).strip().lower()
            if rec_applied == "yes":
                existing["Applied Status"] = "Yes"
                existing["Applied"] = "Yes"

            # 2. Preserve Notes
            if rec.get("Notes") and not existing.get("Notes"):
                existing["Notes"] = rec["Notes"]

            # 3. Preserve earliest Crawl Date
            existing_date = str(existing.get("Crawl Date", existing.get("Date", "")))
            rec_date = str(rec.get("Crawl Date", rec.get("Date", "")))
            if rec_date and (not existing_date or rec_date < existing_date):
                existing["Crawl Date"] = rec_date
                existing["Date"] = rec_date

            # 4. Fill in missing fields
            for k, v in rec.items():
                if v and (not existing.get(k) or str(existing.get(k)).strip() in {"", "N/A", "Unknown", "-"}):
                    existing[k] = v
        else:
            unique_map[key] = rec

    return list(unique_map.values()), dupes_count


def clean_excel_file(file_path: Path, dry_run: bool = False) -> int:
    """Read an Excel workbook, deduplicate all rows based on LinkedIn URL, and re-save."""
    if not file_path.exists():
        log(f"File not found: {file_path}", "ERROR")
        return 0

    log(f"Scanning Excel file: {file_path}", "INFO")
    
    # Read all sheets using pandas
    xls = pd.ExcelFile(file_path)
    sheet_names = xls.sheet_names
    total_file_dupes = 0
    cleaned_sheets = {}

    for sname in sheet_names:
        df = pd.read_excel(xls, sheet_name=sname)
        if df.empty:
            cleaned_sheets[sname] = df
            continue

        records = df.to_dict(orient="records")
        deduped, dupes = deduplicate_records(records)
        total_file_dupes += dupes

        clean_df = pd.DataFrame(deduped)
        # Re-number the '#' column if it exists
        if "#" in clean_df.columns:
            clean_df["#"] = range(1, len(clean_df) + 1)

        cleaned_sheets[sname] = clean_df
        log(f"   Sheet [{sname}]: {len(records)} rows -> {len(clean_df)} unique rows ({dupes} duplicates removed)")

    if total_file_dupes == 0:
        log(f"✅ Excel file is already 100% clean (0 duplicates found).", "OK")
        return 0

    if dry_run:
        log(f"DRY RUN: {total_file_dupes} total duplicates identified (no files modified).", "WARN")
        return total_file_dupes

    # Save cleaned workbook with professional styling
    log(f"Writing cleaned Excel workbook ({total_file_dupes} duplicates purged) …", "INFO")
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    cols = ["#","Crawl Date","Country","Flag","Tier","City","City Openings",
            "Location","Company","Job Title","Search Keyword","Posted Date",
            "Easy Apply","Remote / Workplace","Match Score","Visa Sponsorship",
            "Skills","Resume","Applied","Notes","Job URL"]

    hf   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hfnt = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    vf   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for sname, df in cleaned_sheets.items():
        ws = wb.create_sheet(title=sname[:31])
        if df.empty:
            continue

        # Header
        sheet_cols = list(df.columns)
        for ci, col in enumerate(sheet_cols, 1):
            c = ws.cell(row=1, column=ci, value=str(col))
            c.fill = hf; c.font = hfnt
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin

        # Rows
        for ri, (_, row) in enumerate(df.iterrows(), 2):
            for ci, col in enumerate(sheet_cols, 1):
                val = row.get(col, "")
                if pd.isna(val):
                    val = ""
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Calibri", size=9); c.border = thin
                if str(col).lower() in {"visa sponsorship", "visa sponsorship mentioned"} and str(val).lower() == "yes":
                    c.fill = vf

        for i in range(1, len(sheet_cols) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        ws.freeze_panes = "A2"
        if len(df) > 0:
            ws.auto_filter.ref = ws.dimensions

    try:
        wb.save(file_path)
        log(f"✅ Cleaned Excel saved -> {file_path}", "OK")
    except PermissionError:
        alt_path = file_path.parent / f"{file_path.stem}_cleaned_{int(time.time())}.xlsx"
        wb.save(alt_path)
        log(f"Original file locked — saved to {alt_path}", "WARN")

    return total_file_dupes


def clean_csv_file(file_path: Path, dry_run: bool = False) -> int:
    """Read a CSV file, deduplicate based on LinkedIn URL, and re-save."""
    if not file_path.exists():
        return 0

    log(f"Scanning CSV file: {file_path}", "INFO")
    df = pd.read_csv(file_path, on_bad_lines="skip")
    if df.empty:
        return 0

    records = df.to_dict(orient="records")
    deduped, dupes = deduplicate_records(records)
    log(f"   CSV: {len(records)} rows -> {len(deduped)} unique rows ({dupes} duplicates removed)")

    if dupes > 0 and not dry_run:
        clean_df = pd.DataFrame(deduped)
        clean_df.to_csv(file_path, index=False)
        log(f"✅ Cleaned CSV saved -> {file_path}", "OK")

    return dupes


def main():
    parser = argparse.ArgumentParser(description="Remove duplicate jobs from Excel/CSV based on LinkedIn URL")
    parser.add_argument("--file", type=str, default=None,
                        help="Specific Excel or CSV file to deduplicate")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview duplicate removal without modifying files")
    args = parser.parse_args()

    log("=" * 65)
    log("🚀  LinkedIn Job Deduplicator (Excel & CSV)")
    log("=" * 65)

    if args.file:
        p = Path(args.file)
        if p.suffix.lower() in {".xlsx", ".xls"}:
            clean_excel_file(p, dry_run=args.dry_run)
        elif p.suffix.lower() == ".csv":
            clean_csv_file(p, dry_run=args.dry_run)
        else:
            log(f"Unsupported file format: {p.suffix}", "ERROR")
    else:
        # Default: clean master project files
        if MASTER_EXCEL_PATH.exists():
            clean_excel_file(MASTER_EXCEL_PATH, dry_run=args.dry_run)
        if MASTER_CSV_PATH.exists():
            clean_csv_file(MASTER_CSV_PATH, dry_run=args.dry_run)

    log("\n" + "=" * 65)
    log("✅  Deduplication Check Complete", "OK")
    log("=" * 65)


if __name__ == "__main__":
    main()
